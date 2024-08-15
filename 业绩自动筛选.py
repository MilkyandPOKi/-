import pandas as pd
from openpyxl import load_workbook
import os
# 获取当前目录下的所有文件
files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith(('.xlsx', '.xlsm'))]
# 检查是否只有一个xlsx文件
if len(files) != 1:
    print("当前目录下应该有且仅有一个.xlsx文件，但找到了{}个文件。".format(len(files)))
    # 可以选择退出程序或进行其他处理
    exit(1)
# 取出唯一的xlsx文件
file_path = files[0]

#使用pandas读取数据，然后应用筛选规则
df = pd.read_excel(file_path)
filtered_df = df[(df['***办事处'] == '***业务') & (df['***业绩组'] == '***垂管组')]
# 创建一个pandas Excel writer，使用openpyxl引擎，并加载现有的工作簿
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
    filtered_df.to_excel(writer, sheet_name='Filtered_Data', index=False)

# 使用pandas读取数据并应用筛选规则
#读取文件
df1 = pd.read_excel(file_path)
#定义筛选的组，其中excluded_people是要进行反选的
filter_groups = ['XXX组', 'XXX组', 'XXX组', 'XXX组', 'XXX组', 'XXX组']
excluded_people = ['XXX', 'XXX', 'XXX', 'XXX']
#应用筛选条件
filtered_df_1 = df1[df1['***业绩组'].isin(filter_groups)]
filtered_df_2 = filtered_df_1[~filtered_df_1['***业绩人'].isin(excluded_people)]

wb = load_workbook(file_path)
ws = wb['Filtered_Data']
#找到最后一行
last_row = ws.max_row

# 由于我们不包含表头，所以从最后一行的下一行（即last_row + 1）开始写入数据
for r_idx, row in enumerate(filtered_df_2.itertuples(index=False), start=last_row + 1):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

    # 保存修改后的工作簿
wb.save(file_path)